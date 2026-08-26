"""
Generates a downloadable Excel template with flat single-row headers
and one pre-filled sample row showing correct format for all 50 fields.
"""
import io
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from app.services.import_schema import FIELD_HEADERS

HEADERS = [header_text for _field_key, header_text in FIELD_HEADERS]

SAMPLE_ROW = [
    # General
    1, "LAR", "GEC-II", "GEC-II Guwahati", "Guwahati",
    "Sample 33/11kV SS", "26.1839", "91.6667", "Conventional",
    # GSS
    "132kV Sishugram GSS", "",
    "", "",
    # Transformer
    2.5, "BHEL", 2010,
    1.8, 75.0, 65.0,
    # Feeder (first feeder row for this substation)
    "33kV Ulubari Incomer", "Substation Incomer", "33kV",
    # Meter
    "MTR001234", "L&T", "DLMS",
    "Working", "100/5A", 6000.0,
    "Panel Mounted", "Working",
    "Panel Mounted", "Working",
    "Working",
    # Switchgear
    "Indoor", "Crompton Greaves", "Working",
    "CGL", 2012,
    "Numerical",
    "None",
    "ABB REF615", "",
    "Working", "Working", "Working",
    2012,
    "",
    # DC Supply
    "Working", "Amararaja", 2015,
    "Working", "VRLA",
]

# Colour scheme
CLR_HEADER_GENERAL  = "1A2744"  # dark blue — general info
CLR_HEADER_GSS      = "4A235A"  # purple — GSS
CLR_HEADER_TR       = "7B1C1C"  # dark red — transformer
CLR_HEADER_FEEDER   = "1A4A7B"  # dark blue — feeder
CLR_HEADER_METER    = "1A5C2A"  # dark green — meter
CLR_HEADER_SWG      = "7B4A00"  # dark amber — switchgear
CLR_HEADER_DC       = "2A4A4A"  # dark teal — DC supply
CLR_SAMPLE          = "FFF9E6"  # light yellow — sample row

# Column ranges for each group (1-based)
GROUP_COLORS = {
    range(1, 10):  CLR_HEADER_GENERAL,
    range(10, 14): CLR_HEADER_GSS,
    range(14, 20): CLR_HEADER_TR,
    range(20, 23): CLR_HEADER_FEEDER,
    range(23, 34): CLR_HEADER_METER,
    range(34, 48): CLR_HEADER_SWG,
    range(48, 53): CLR_HEADER_DC,
}


def get_header_color(col_idx_1based):
    for rng, clr in GROUP_COLORS.items():
        if col_idx_1based in rng:
            return clr
    return "333333"


def generate_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Substation Data"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Write headers ──
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        clr  = get_header_color(col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=9,
                              name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=clr)
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="center")
        cell.border    = border
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max(14, len(header) * 0.9)

    ws.row_dimensions[1].height = 40

    # ── Write sample row ──
    for col_idx, value in enumerate(SAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill      = PatternFill("solid", fgColor=CLR_SAMPLE)
        cell.font      = Font(size=9, italic=True, name="Calibri")
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border    = border

    ws.row_dimensions[2].height = 18

    # Freeze panes — keep header and first 2 cols visible
    ws.freeze_panes = "C2"

    # ── Instructions sheet ──
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["APDCL Sub-Station Monitoring Portal — Excel Import Template"],
        [""],
        ["IMPORTANT RULES:"],
        ["1. Do NOT modify the header row (Row 1)."],
        ["2. Delete the sample row (Row 2) before importing."],
        ["3. One row per FEEDER. Transformer and substation details repeat on the first feeder row of each substation block."],
        ["4. Leave cells blank (not zero) when data is not available."],
        ["5. Latitude/Longitude can be decimal (26.1839) or DMS format (26° 11' 2.09\"N)."],
        ["6. Feeder Voltage must be '33kV' or '11kV' exactly."],
        ["7. Feeder Type: use 'Substation Incomer', 'Transformer Incomer', 'Transformer Outgoing' or 'Outgoing Feeder'."],
        ["8. VCB Type: use 'Indoor' or 'Outdoor' (or 'Autorecloser' — see rule 11)."],
        ["9. Meter Type: use 'DLMS' or 'Non-DLMS'."],
        ["10. Substations are matched by Substation Name + ESD + Primary GSS together. If two substations share a name, make sure their ESD or Primary GSS differs so they don't get merged."],
        ["11. For autorecloser feeders: set VCB Type to 'Autorecloser' — the system will detect and render the AR symbol automatically."],
        ["12. Bus coupler rows: set Feeder Name to 'Bus Coupler' (Feeder Type is ignored for these rows — they are always classified as a bus coupler and shown in the topology, not as a regular feeder)."],
        [""],
        ["COLUMN GROUPS (colour coded):"],
        ["Dark Blue (cols 1-9)  — General substation details"],
        ["Purple (cols 10-13)   — GSS connectivity"],
        ["Dark Red (cols 14-19) — Power transformer details (repeat for each transformer)"],
        ["Blue (cols 20-22)     — Feeder name, type and voltage"],
        ["Green (cols 23-33)    — Meter, CT and PT details"],
        ["Amber (cols 34-47)    — Switchgear and protection relay details"],
        ["Teal (cols 48-52)     — DC supply (battery charger and battery bank)"],
    ]
    for i, row in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=row[0])
        if i == 1:
            cell.font = Font(bold=True, size=13, color="1A2744", name="Calibri")
        elif row[0].startswith(("IMPORTANT", "COLUMN")):
            cell.font = Font(bold=True, size=10, color="CC2200", name="Calibri")
        else:
            cell.font = Font(size=9, name="Calibri")
    ws2.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
